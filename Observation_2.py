import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import re
import os
import uuid
import time
import threading

class TCPPortAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("TCP Port Analysis Application")
        self.root.geometry("1200x800")
        
        # Initialize data storage
        self.annexure_data = None
        self.template_data = None
        self.filtered_data = None
        self.defined_data = None
        self.matched_data = None
        self.not_found_data = None
        
        # Define observation mappings
        self.ob_name_mapping = {
            'vnc': "Older Version of VNC (Virtual Network Computing) Running",
            'vnc-http': "Older Version of VNC (Virtual Network Computing) Running",
            'telnet': "Unencrypted Telnet Server",
            'unknown': "Unknown Port",
            'ftp': "Clear Text Protocol FTP Open Port",
            'http': "Clear Text Protocol HTTP Open Port",
            'ssh': "Older Version of OpenSSH",
            'dropbear': "Dropbear sshd 2018.76 (protocol 2.0)"
        }
        
        # Apply dark theme
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('TFrame', background='#2b2b2b')
        self.style.configure('TButton', background='#3c3f41', foreground='white')
        self.style.configure('TLabel', background='#2b2b2b', foreground='white')
        self.style.configure('Treeview', 
                           background='#2b2b2b',
                           foreground='white',
                           fieldbackground='#2b2b2b')
        self.style.configure('Treeview.Heading', 
                           background='#3c3f41',
                           foreground='white')
        self.style.configure('TProgressbar', 
                           background='#00B050',
                           troughcolor='#3c3f41')
        self.root.configure(bg='#2b2b2b')
        
        self.setup_ui()
    
    def setup_ui(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        nav_frame = ttk.Frame(main_frame)
        nav_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.upload_annexure_btn = ttk.Button(nav_frame, text="Upload Annexure 2", 
                                            command=self.upload_annexure)
        self.upload_annexure_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.upload_template_btn = ttk.Button(nav_frame, text="Upload Observation Template", 
                                            command=self.upload_template)
        self.upload_template_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.process_btn = ttk.Button(main_frame, text="Process", 
                                    command=self.start_process_thread, state='disabled')
        self.process_btn.grid(row=1, column=0, pady=(0, 10), sticky=tk.W)
        
        self.export_btn = ttk.Button(main_frame, text="Export", 
                                   command=self.export_data, state='disabled')
        self.export_btn.grid(row=1, column=1, pady=(0, 10), sticky=tk.E)
        
        self.progress_bar = ttk.Progressbar(main_frame, mode='determinate', length=400)
        self.progress_bar.grid(row=1, column=0, columnspan=2, pady=(0, 10), sticky=tk.EW)
        self.progress_bar.grid_remove()
        
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.tab1_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab1_frame, text="Filtered Data")
        self.setup_tab1()
        
        self.tab2_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab2_frame, text="Defined Observations")
        self.setup_tab2()
        
        self.tab3_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.tab3_frame, text="Matched Templates")
        self.setup_tab3()
    
    def setup_tab1(self):
        self.tree1 = ttk.Treeview(self.tab1_frame, 
                                columns=('IP', 'Port', 'Status', 'Service', 'Version'), 
                                show='headings', height=15)
        
        column_titles = {
            'IP': 'IP Address',
            'Port': 'Port Number',
            'Status': 'Status',
            'Service': 'Service',
            'Version': 'Version'
        }
        
        for col, title in column_titles.items():
            self.tree1.heading(col, text=title)
        
        self.tree1.column('IP', width=150, anchor='center')
        self.tree1.column('Port', width=120, anchor='center')
        self.tree1.column('Status', width=100, anchor='center')
        self.tree1.column('Service', width=120, anchor='center')
        self.tree1.column('Version', width=250, anchor='center')
        
        v_scrollbar1 = ttk.Scrollbar(self.tab1_frame, orient="vertical", command=self.tree1.yview)
        h_scrollbar1 = ttk.Scrollbar(self.tab1_frame, orient="horizontal", command=self.tree1.xview)
        self.tree1.configure(yscrollcommand=v_scrollbar1.set, xscrollcommand=h_scrollbar1.set)
        
        self.tree1.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar1.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar1.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.stats_label1 = ttk.Label(self.tab1_frame, 
                                    text="Total Rows: 0, Total Columns: 5", 
                                    background='#90EE90', 
                                    foreground='black')
        self.stats_label1.grid(row=2, column=0, columnspan=2, pady=5, sticky=(tk.W, tk.E))
        
        self.tab1_frame.columnconfigure(0, weight=1)
        self.tab1_frame.rowconfigure(0, weight=1)
    
    def setup_tab2(self):
        self.tree2 = ttk.Treeview(self.tab2_frame, 
                                columns=('IP', 'Port', 'Status', 'Service', 'Version', 'ObsName'), 
                                show='headings', height=15)
        
        column_titles = {
            'IP': 'IP Address',
            'Port': 'Port Number',
            'Status': 'Status',
            'Service': 'Service',
            'Version': 'Version',
            'ObsName': 'Defined Observation Name'
        }
        
        for col, title in column_titles.items():
            self.tree2.heading(col, text=title)
        
        self.tree2.column('IP', width=150, anchor='center')
        self.tree2.column('Port', width=120, anchor='center')
        self.tree2.column('Status', width=100, anchor='center')
        self.tree2.column('Service', width=120, anchor='center')
        self.tree2.column('Version', width=250, anchor='center')
        self.tree2.column('ObsName', width=350, anchor='center')
        
        v_scrollbar2 = ttk.Scrollbar(self.tab2_frame, orient="vertical", command=self.tree2.yview)
        h_scrollbar2 = ttk.Scrollbar(self.tab2_frame, orient="horizontal", command=self.tree2.xview)
        self.tree2.configure(yscrollcommand=v_scrollbar2.set, xscrollcommand=h_scrollbar2.set)
        
        self.tree2.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar2.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar2.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.stats_label2 = ttk.Label(self.tab2_frame, 
                                    text="Total Rows: 0, Total Columns: 6", 
                                    background='#90EE90', 
                                    foreground='black')
        self.stats_label2.grid(row=2, column=0, columnspan=2, pady=5, sticky=(tk.W, tk.E))
        
        self.tab2_frame.columnconfigure(0, weight=1)
        self.tab2_frame.rowconfigure(0, weight=1)
    
    def setup_tab3(self):
        columns = ['S_No', 'Category', 'Observation Name', 'Affected IP Address', 
                  'Port Number', 'Port Type', 'Device Type', 'Hostname', 
                  'Device Location', 'Observation Description', 'Severity', 
                  'Risk', 'Recommendation', 'Patch Priority', 'Patch Status']
        
        self.tree3 = ttk.Treeview(self.tab3_frame, columns=columns, show='headings', height=15)
        
        for col in columns:
            self.tree3.heading(col, text=col)
            self.tree3.column(col, width=120, anchor='center')
        
        v_scrollbar3 = ttk.Scrollbar(self.tab3_frame, orient="vertical", command=self.tree3.yview)
        h_scrollbar3 = ttk.Scrollbar(self.tab3_frame, orient="horizontal", command=self.tree3.xview)
        self.tree3.configure(yscrollcommand=v_scrollbar3.set, xscrollcommand=h_scrollbar3.set)
        
        self.tree3.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar3.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar3.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        self.stats_label3 = ttk.Label(self.tab3_frame, 
                                    text="Total Rows: 0, Total Columns: 15", 
                                    background='#90EE90', 
                                    foreground='black')
        self.stats_label3.grid(row=2, column=0, columnspan=2, pady=5, sticky=(tk.W, tk.E))
        
        self.tab3_frame.columnconfigure(0, weight=1)
        self.tab3_frame.rowconfigure(0, weight=1)
    
    def upload_annexure(self):
        file_path = filedialog.askopenfilename(
            title="Select Annexure 2 File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")]
        )
        
        if file_path:
            try:
                if file_path.endswith('.csv'):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)
                
                tcp_header_row = self.find_tcp_header(df)
                if tcp_header_row is not None:
                    self.annexure_data = self.extract_tcp_data(df, tcp_header_row)
                    self.check_enable_process()
                else:
                    messagebox.showerror("Error", "TCP table header not found in the file!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to upload Annexure 2: {str(e)}")
    
    def find_tcp_header(self, df):
        target_columns = ['IP Address', 'Port Number', 'Status', 'Service', 'Version']
        
        for row_idx in range(len(df)):
            row_data = df.iloc[row_idx].astype(str).str.lower()
            found_columns = 0
            
            for col in target_columns:
                if any(col.lower() in cell for cell in row_data if pd.notna(cell)):
                    found_columns += 1
            
            if found_columns >= 3:
                return row_idx
        
        return None
    
    def extract_tcp_data(self, df, header_row):
        headers = df.iloc[header_row].astype(str)
        data_rows = df.iloc[header_row + 1:].reset_index(drop=True)
        
        ip_col = port_col = status_col = service_col = version_col = None
        
        for idx, header in enumerate(headers):
            header_lower = str(header).lower()
            if 'ip address' in header_lower or 'ip' in header_lower:
                ip_col = idx
            elif 'port number' in header_lower or 'port' in header_lower:
                port_col = idx
            elif 'status' in header_lower:
                status_col = idx
            elif 'service' in header_lower:
                service_col = idx
            elif 'version' in header_lower:
                version_col = idx
        
        tcp_data = []
        for _, row in data_rows.iterrows():
            if ip_col is not None and pd.notna(row.iloc[ip_col]):
                port_str = str(row.iloc[port_col]) if port_col is not None else ''
                port_num = re.search(r'(\d+)', port_str)
                port_value = int(port_num.group(1)) if port_num else 0
                service = str(row.iloc[service_col]).lower() if service_col is not None else ''
                
                if 'ssl/unknown' in service or 'ssl/http' in service:
                    continue
                    
                tcp_data.append({
                    'IP Address': str(row.iloc[ip_col]) if ip_col is not None else '',
                    'Port Number': port_value,
                    'Status': str(row.iloc[status_col]) if status_col is not None else '',
                    'Service': service,
                    'Version': str(row.iloc[version_col]).lower() if version_col is not None else ''
                })
        
        return pd.DataFrame(tcp_data)
    
    def upload_template(self):
        file_path = filedialog.askopenfilename(
            title="Select Observation Template File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")]
        )
        
        if file_path:
            try:
                if file_path.endswith('.csv'):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)
                
                template_header_row = self.find_template_header(df)
                if template_header_row is not None:
                    self.template_data = self.extract_template_data(df, template_header_row)
                    self.check_enable_process()
                else:
                    messagebox.showerror("Error", "Template header not found in the file!")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to upload template: {str(e)}")
    
    def find_template_header(self, df):
        target_columns = ['S_No', 'Category', 'Observation Name', 'Affected IP Address', 
                         'Port Number', 'Port Type', 'Device Type']
        
        for row_idx in range(len(df)):
            row_data = df.iloc[row_idx].astype(str).str.lower()
            found_columns = 0
            
            for col in target_columns:
                if any(col.lower() in cell for cell in row_data if pd.notna(cell)):
                    found_columns += 1
            
            if found_columns >= 4:
                return row_idx
        
        return None
    
    def extract_template_data(self, df, header_row):
        headers = df.iloc[header_row]
        data_rows = df.iloc[header_row + 1:].reset_index(drop=True)
        
        template_data = data_rows.copy()
        template_data.columns = headers
        
        return template_data
    
    def check_enable_process(self):
        if self.annexure_data is not None and self.template_data is not None:
            self.process_btn.config(state='normal')
    
    def start_process_thread(self):
        self.process_btn.config(state='disabled')
        self.progress_bar.grid()
        self.progress_bar['value'] = 0
        
        def process_with_progress():
            try:
                steps = 4
                for i in range(steps):
                    if i == 0:
                        self.filter_annexure_data()
                    elif i == 1:
                        self.add_observation_names()
                    elif i == 2:
                        self.match_with_template()
                    elif i == 3:
                        self.update_ui_tabs()
                    self.progress_bar['value'] = ((i + 1) / steps) * 100
                    self.root.update()
                    time.sleep(0.5)
                
                self.export_btn.config(state='normal')
            except Exception as e:
                messagebox.showerror("Error", f"Failed to process data: {str(e)}")
            finally:
                self.progress_bar.grid_remove()
                self.process_btn.config(state='normal')
        
        threading.Thread(target=process_with_progress, daemon=True).start()
    
    def filter_annexure_data(self):
        filtered_rows = []
        
        for _, row in self.annexure_data.iterrows():
            service = row['Service'].lower()
            version = row['Version'].lower()
            
            if self.matches_service_criteria(service, version):
                filtered_rows.append(row)
        
        self.filtered_data = pd.DataFrame(filtered_rows)
    
    def matches_service_criteria(self, service, version):
        service_filters = ['vnc', 'vnc-http', 'telnet', 'unknown', 'ftp', 'http', 'ssh', 'dropbear']
        
        if 'odette-ftp' in service:
            return False
            
        for filter_service in service_filters:
            if filter_service == 'http':
                if service == 'http' and ('http' in version.lower()):
                    return True
            elif filter_service == 'ssh':
                if service == 'ssh' and 'openssh' in version:
                    return True
            elif filter_service == 'dropbear':
                if 'dropbear' in version:
                    return True
            elif filter_service == 'ftp':
                if service == 'ftp':
                    return True
            elif filter_service in service:
                return True
        
        return False
    
    def add_observation_names(self):
        defined_rows = []
        
        for _, row in self.filtered_data.iterrows():
            service = row['Service'].lower()
            version = row['Version'].lower()
            
            obs_name = self.get_observation_name(service, version)
            
            new_row = row.to_dict()
            new_row['Defined Observation Name'] = obs_name
            defined_rows.append(new_row)
        
        self.defined_data = pd.DataFrame(defined_rows)
    
    def get_observation_name(self, service, version):
        if 'vnc-http' in service:
            return self.ob_name_mapping['vnc-http']
        elif 'vnc' in service:
            return self.ob_name_mapping['vnc']
        elif 'telnet' in service:
            return self.ob_name_mapping['telnet']
        elif 'unknown' in service:
            return self.ob_name_mapping['unknown']
        elif service == 'ftp':
            return self.ob_name_mapping['ftp']
        elif 'dropbear' in version:
            return self.ob_name_mapping['dropbear']
        elif service == 'http' and 'http' in version.lower():
            return self.ob_name_mapping['http']
        elif service == 'ssh' and 'openssh' in version:
            return self.ob_name_mapping['ssh']
        return "Unknown Service"
    
    def match_with_template(self):
        matched_rows = []
        not_found_rows = []
        
        for idx, defined_row in self.defined_data.iterrows():
            obs_name = defined_row['Defined Observation Name']
            
            template_match = None
            for _, template_row in self.template_data.iterrows():
                template_obs_name = str(template_row.get('Observation Name', '')).strip().lower()
                if obs_name.lower() == template_obs_name or \
                   (defined_row['Service'] == 'http' and template_obs_name == self.ob_name_mapping['http'].lower()):
                    template_match = template_row
                    break
            
            port_num = defined_row['Port Number']
            if isinstance(port_num, str):
                port_num = re.search(r'(\d+)', port_num)
                port_num = port_num.group(1) if port_num else ''
            
            if template_match is not None:
                matched_row = {
                    'S_No': idx + 1,
                    'Category': template_match.get('Category', ''),
                    'Observation Name': template_match.get('Observation Name', ''),
                    'Affected IP Address': defined_row['IP Address'],
                    'Port Number': port_num,
                    'Port Type': 'TCP',
                    'Device Type': template_match.get('Device Type', ''),
                    'Hostname': template_match.get('Hostname / Serial No. / Operating System / Mac Address', ''),
                    'Device Location': template_match.get('Device Location', ''),
                    'Observation Description': template_match.get('Observation Description', ''),
                    'Severity': template_match.get('Severity', ''),
                    'Risk': template_match.get('Risk', ''),
                    'Recommendation': template_match.get('Recommendation', ''),
                    'Patch Priority': template_match.get('Patch Priority', ''),
                    'Patch Status': template_match.get('Patch Status', '')
                }
                matched_rows.append(matched_row)
            elif defined_row['Service'] != 'http':
                not_found_row = defined_row.to_dict()
                not_found_rows.append(not_found_row)
        
        if matched_rows:
            severity_order = {'critical': 1, 'high': 2, 'medium': 3, 'low': 4, 'info': 5}
            matched_rows.sort(key=lambda x: severity_order.get(str(x.get('Severity', '')).lower(), 6))
        
        self.matched_data = pd.DataFrame(matched_rows)
        self.not_found_data = pd.DataFrame(not_found_rows)
    
    def update_ui_tabs(self):
        for tree in [self.tree1, self.tree2, self.tree3]:
            for item in tree.get_children():
                tree.delete(item)
        
        if self.filtered_data is not None:
            for _, row in self.filtered_data.iterrows():
                self.tree1.insert('', 'end', values=(
                    row['IP Address'], row['Port Number'], row['Status'],
                    row['Service'], row['Version']
                ))
            self.stats_label1.config(text=f"Total Rows: {len(self.filtered_data)}, Total Columns: 5")
        
        if self.defined_data is not None:
            for _, row in self.defined_data.iterrows():
                self.tree2.insert('', 'end', values=(
                    row['IP Address'], row['Port Number'], row['Status'],
                    row['Service'], row['Version'], row['Defined Observation Name']
                ))
            self.stats_label2.config(text=f"Total Rows: {len(self.defined_data)}, Total Columns: 6")
        
        if self.matched_data is not None:
            columns = ['S_No', 'Category', 'Observation Name', 'Affected IP Address', 
                      'Port Number', 'Port Type', 'Device Type', 'Hostname', 
                      'Device Location', 'Observation Description', 'Severity', 
                      'Risk', 'Recommendation', 'Patch Priority', 'Patch Status']
            
            for _, row in self.matched_data.iterrows():
                values = [row.get(col, '') for col in columns]
                self.tree3.insert('', 'end', values=values)
            self.stats_label3.config(text=f"Total Rows: {len(self.matched_data)}, Total Columns: 15")
    
    def export_data(self):
        if self.matched_data is None:
            messagebox.showerror("Error", "No data to export!")
            return
        
        file_path = filedialog.asksaveasfilename(
            title="Save Export File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        if file_path:
            try:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    if not self.matched_data.empty:
                        export_columns = ['S_No', 'Category', 'Observation Name', 'Affected IP Address', 
                                       'Port Number', 'Port Type', 'Device Type', 'Hostname', 
                                       'Device Location', 'Observation Description', 'Severity', 
                                       'Risk', 'Recommendation', 'Patch Priority', 'Patch Status']
                        self.matched_data[export_columns].to_excel(writer, sheet_name='Observation 2', index=False)
                    
                    if self.defined_data is not None and not self.defined_data.empty:
                        self.defined_data.to_excel(writer, sheet_name='Filtered Data', index=False)
                    
                    if self.not_found_data is not None and not self.not_found_data.empty:
                        self.not_found_data.to_excel(writer, sheet_name='Not Found', index=False)
                
                self.apply_excel_formatting(file_path)
                messagebox.showinfo("Success", "Data exported successfully!")
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to export data: {str(e)}")
    
    def apply_excel_formatting(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            
            header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            severity_colors = {
                'critical': PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
                'high': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
                'medium': PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
                'low': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
                'info': PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
            }
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            bold_font = Font(bold=True)
            normal_font = Font(bold=False)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                column_widths = {}
                for col in range(1, ws.max_column + 1):
                    max_length = 0
                    column = get_column_letter(col)
                    for row in range(1, ws.max_row + 1):
                        try:
                            cell_value = str(ws[f'{column}{row}'].value)
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except:
                            pass
                    column_widths[column] = min(max_length + 2, 50)
                
                severity_col = None
                if 'Observation 2' in sheet_name:
                    for idx, cell in enumerate(ws[1], 1):
                        if cell.value and 'severity' in str(cell.value).lower():
                            severity_col = idx
                            break
                
                for row in range(1, ws.max_row + 1):
                    ws.row_dimensions[row].height = 30
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = center_align
                        
                        if row == 1:
                            cell.fill = header_fill
                            cell.font = bold_font
                        
                        if severity_col and col == severity_col:
                            cell.font = bold_font
                            if row > 1 and cell.value:
                                severity_value = str(cell.value).lower()
                                if severity_value in severity_colors:
                                    cell.fill = severity_colors[severity_value]
                        else:
                            cell.font = normal_font
                
                for col, width in column_widths.items():
                    ws.column_dimensions[col].width = width * 1.2
            
            wb.save(file_path)
            
        except Exception as e:
            print(f"Warning: Could not apply formatting: {str(e)}")

def main():
    root = tk.Tk()
    app = TCPPortAnalyzer(root)
    root.mainloop()

if __name__ == "__main__":
    main()