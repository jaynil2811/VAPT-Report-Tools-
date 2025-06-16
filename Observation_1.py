import tkinter as tk
from tkinter import filedialog, ttk
import tkinter.font as tkfont
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import re

class NessusObservationProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Nessus Observation Processor")
        self.root.geometry("1000x700")
        self.root.configure(bg="#1E2526")
        self.root.resizable(True, True)

        self.nessus_files = []
        self.template_file = ""
        self.output_df = None
        self.filtered_df = None
        self.not_found_df = None

        # Custom style for dark theme
        self.style = ttk.Style()
        self.style.theme_use("clam")
        self.style.configure("TButton", 
                            background="#00C4B4", 
                            foreground="white", 
                            padding=12, 
                            font=("Roboto", 11, "bold") if "Roboto" in tkfont.families() else ("Segoe UI", 11, "bold"),
                            borderwidth=0,
                            relief="flat")
        self.style.map("TButton", 
                      background=[('active', '#26D9CA')],
                      foreground=[('active', 'white')])
        self.style.configure("TLabel", 
                            background="#1E2526", 
                            foreground="#E0E6E6", 
                            font=("Roboto", 10) if "Roboto" in tkfont.families() else ("Segoe UI", 10))
        self.style.configure("TProgressbar", 
                            background="#00C4B4", 
                            troughcolor="#2A3233",
                            thickness=20,
                            borderwidth=0)
        self.style.configure("TFrame", 
                            background="#1E2526")
        self.style.configure("TNotebook", 
                            background="#1E2526", 
                            foreground="#E0E6E6")
        self.style.configure("TNotebook.Tab", 
                            background="#2A3233", 
                            foreground="#E0E6E6", 
                            padding=[15, 8],
                            font=("Roboto", 10, "bold") if "Roboto" in tkfont.families() else ("Segoe UI", 10, "bold"),
                            borderwidth=0)
        self.style.map("TNotebook.Tab", 
                      background=[('selected', '#00C4B4')],
                      foreground=[('selected', 'white')])
        self.style.configure("Treeview", 
                            background="#2A3233", 
                            foreground="#E0E6E6", 
                            fieldbackground="#2A3233",
                            font=("Roboto", 10) if "Roboto" in tkfont.families() else ("Segoe UI", 10),
                            rowheight=35)
        self.style.configure("Treeview.Heading", 
                            background="#00C4B4", 
                            foreground="white",
                            font=("Roboto", 10, "bold") if "Roboto" in tkfont.families() else ("Segoe UI", 10, "bold"),
                            borderwidth=0)
        self.style.map("Treeview", 
                      background=[('!selected', '#2A3233'), ('selected', '#00C4B4')],
                      foreground=[('!selected', '#E0E6E6'), ('selected', 'white')])
        self.style.layout("Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        # Create a canvas and scrollbar for the main UI
        self.canvas = tk.Canvas(self.root, bg="#1E2526", highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True, padx=20, pady=20)
        self.scrollbar.pack(side="right", fill="y")

        self.canvas_frame = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Main frame (inside the scrollable frame)
        self.main_frame = ttk.Frame(self.scrollable_frame)
        self.main_frame.pack(fill="both", expand=True, padx=15, pady=15)
        self.main_frame.columnconfigure(0, weight=1)
        self.main_frame.rowconfigure(5, weight=1)

        # Header (Centered)
        self.header_label = ttk.Label(self.main_frame, 
                                     text="Nessus Observation Processor",
                                     font=("Roboto", 24, "bold") if "Roboto" in tkfont.families() else ("Segoe UI", 24, "bold"),
                                     foreground="#00C4B4")
        self.header_label.grid(row=0, column=0, pady=(0, 20), sticky="ew")

        # File selection section (with border)
        self.file_section_frame = ttk.Frame(self.main_frame, style="TFrame")
        self.file_section_frame.grid(row=1, column=0, pady=10, sticky="ew")
        self.file_section_frame.configure(style="Border.TFrame")
        self.style.configure("Border.TFrame", background="#2A3233", relief="solid", borderwidth=1)

        self.nessus_frame = ttk.Frame(self.file_section_frame)
        self.nessus_frame.pack(fill="x", padx=15, pady=10)
        self.nessus_label = ttk.Label(self.nessus_frame, 
                                     text="Nessus CSV: Not selected",
                                     wraplength=600)
        self.nessus_label.pack(side="left", padx=10)
        self.nessus_btn = ttk.Button(self.nessus_frame, 
                                    text="Upload CSV", 
                                    command=self.upload_nessus)
        self.nessus_btn.pack(side="right")

        self.template_frame = ttk.Frame(self.file_section_frame)
        self.template_frame.pack(fill="x", padx=15, pady=10)
        self.template_label = ttk.Label(self.template_frame, 
                                       text="Observation Template: Not selected",
                                       wraplength=600)
        self.template_label.pack(side="left", padx=10)
        self.template_btn = ttk.Button(self.template_frame, 
                                      text="Upload Excel", 
                                      command=self.upload_template)
        self.template_btn.pack(side="right")

        # Process and Export buttons section (Centered)
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.grid(row=3, column=0, pady=20, sticky="ew")
        self.button_frame.columnconfigure(0, weight=1)
        self.button_frame.columnconfigure(1, weight=1)

        self.process_btn = ttk.Button(self.button_frame, 
                                     text="Process", 
                                     command=self.process_files)
        self.process_btn.grid(row=0, column=0, padx=(0, 10), sticky="e")

        self.export_btn = ttk.Button(self.button_frame, 
                                    text="Export as Excel", 
                                    command=self.export_to_excel)
        self.export_btn.grid(row=0, column=1, padx=(10, 0), sticky="w")
        self.export_btn.state(['disabled'])

        # Progress bar (Centered)
        self.progress = ttk.Progressbar(self.main_frame, 
                                       length=700, 
                                       mode="determinate")
        self.progress.grid(row=4, column=0, pady=20, sticky="ew")

        # Status label (Centered)
        self.status_label = ttk.Label(self.main_frame, 
                                     text="Status: Ready",
                                     font=("Roboto", 10, "italic") if "Roboto" in tkfont.families() else ("Segoe UI", 10, "italic"),
                                     foreground="#E0E6E6")
        self.status_label.grid(row=5, column=0, pady=5, sticky="ew")

        # Notebook for tabs (Stretched)
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.grid(row=6, column=0, pady=10, sticky="nsew")

        # Filtered Data Tab
        self.filtered_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.filtered_tab, text="Filtered Data")
        self.filtered_frame = ttk.Frame(self.filtered_tab)
        self.filtered_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.filtered_frame.columnconfigure(0, weight=1)
        self.filtered_frame.rowconfigure(0, weight=1)
        
        self.filtered_tree = ttk.Treeview(self.filtered_frame, 
                                         columns=("Risk", "Host", "Protocol", "Port", "Name"),
                                         show="headings")
        for col in ("Risk", "Host", "Protocol", "Port", "Name"):
            self.filtered_tree.heading(col, text=col)
            self.filtered_tree.column(col, width=200, stretch=True)
        self.filtered_tree.grid(row=0, column=0, sticky="nsew")
        
        self.filtered_v_scroll = ttk.Scrollbar(self.filtered_frame, 
                                              orient="vertical", 
                                              command=self.filtered_tree.yview)
        self.filtered_v_scroll.grid(row=0, column=1, sticky="ns")
        self.filtered_tree.configure(yscrollcommand=self.filtered_v_scroll.set)
        
        self.filtered_h_scroll = ttk.Scrollbar(self.filtered_tab, 
                                              orient="horizontal", 
                                              command=self.filtered_tree.xview)
        self.filtered_h_scroll.pack(side="bottom", fill="x")
        self.filtered_tree.configure(xscrollcommand=self.filtered_h_scroll.set)
        
        self.filtered_stats = ttk.Label(self.filtered_tab, 
                                       text="Total Rows: 0 | Total Columns: 0 | Total Values: 0",
                                       font=("Roboto", 8, "italic") if "Roboto" in tkfont.families() else ("Segoe UI", 8, "italic"),
                                       foreground="#00C4B4",
                                       background="#2A3233")
        self.filtered_stats.pack(pady=5, fill="x")

        # Bind Select All and Copy
        self.filtered_tree.bind("<Control-a>", self.select_all_filtered)
        self.filtered_tree.bind("<Control-c>", self.copy_filtered)

        # Final Output Tab
        self.output_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.output_tab, text="Final Output")
        self.output_frame = ttk.Frame(self.output_tab)
        self.output_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.output_frame.columnconfigure(0, weight=1)
        self.output_frame.rowconfigure(0, weight=1)
        
        self.output_tree = ttk.Treeview(self.output_frame, 
                                       show="headings")
        self.output_tree.grid(row=0, column=0, sticky="nsew")
        
        self.output_v_scroll = ttk.Scrollbar(self.output_frame, 
                                            orient="vertical", 
                                            command=self.output_tree.yview)
        self.output_v_scroll.grid(row=0, column=1, sticky="ns")
        self.output_tree.configure(yscrollcommand=self.output_v_scroll.set)
        
        self.output_h_scroll = ttk.Scrollbar(self.output_tab, 
                                            orient="horizontal", 
                                            command=self.output_tree.xview)
        self.output_h_scroll.pack(side="bottom", fill="x")
        self.output_tree.configure(xscrollcommand=self.output_h_scroll.set)
        
        self.output_stats = ttk.Label(self.output_tab, 
                                     text="Total Rows: 0 | Total Columns: 0 | Total Values: 0 | Matches: 0",
                                     font=("Roboto", 8, "italic") if "Roboto" in tkfont.families() else ("Segoe UI", 8, "italic"),
                                     foreground="#00C4B4",
                                     background="#2A3233")
        self.output_stats.pack(pady=5, fill="x")

        # Bind Select All and Copy for Output Treeview
        self.output_tree.bind("<Control-a>", self.select_all_output)
        self.output_tree.bind("<Control-c>", self.copy_output)

        # Bind mouse wheel scrolling to the canvas
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)

    def _on_mousewheel(self, event):
        if event.num == 5 or event.delta < 0:
            self.canvas.yview_scroll(1, "units")
        elif event.num == 4 or event.delta > 0:
            self.canvas.yview_scroll(-1, "units")

    def clean_string(self, text):
        text = str(text).lower()
        text = re.sub(r'[^\w\s]', '', text)
        text = re.sub(r'\s+', ' ', text).strip()
        return text

    def select_all_filtered(self, event):
        self.filtered_tree.selection_set(self.filtered_tree.get_children())
        return "break"

    def copy_filtered(self, event):
        selected_items = self.filtered_tree.selection()
        if not selected_items:
            return "break"
        copy_text = []
        for item in selected_items:
            values = self.filtered_tree.item(item, "values")
            copy_text.append("\t".join(str(v) for v in values))
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(copy_text))
        return "break"

    def select_all_output(self, event):
        self.output_tree.selection_set(self.output_tree.get_children())
        return "break"

    def copy_output(self, event):
        selected_items = self.output_tree.selection()
        if not selected_items:
            return "break"
        copy_text = []
        for item in selected_items:
            values = self.output_tree.item(item, "values")
            copy_text.append("\t".join(str(v) for v in values))
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(copy_text))
        return "break"

    def upload_nessus(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        if file_paths:
            self.nessus_files = list(file_paths)
            file_names = [os.path.basename(path) for path in self.nessus_files]
            display_text = "Nessus CSV: " + ", ".join(file_names) if file_names else "Nessus CSV: Not selected"
            self.nessus_label.config(text=display_text)
            self.status_label.config(text=f"Status: Nessus CSV selected: {len(self.nessus_files)} file(s)")
            self.root.update()

    def upload_template(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.template_file = file_path
            file_name = os.path.basename(file_path)
            self.template_label.config(text=f"Observation Template: {file_name}")
            self.status_label.config(text=f"Status: Observation Template selected: {file_name}")
            self.root.update()

    def find_header_row(self, df, header_name="S_No"):
        for i, row in df.iterrows():
            for col in row:
                if isinstance(col, str) and col.strip().lower() == header_name.lower():
                    return i
        return None

    def process_files(self):
        if not self.nessus_files or not self.template_file:
            self.status_label.config(text="Error: Please upload both Nessus CSV(s) and Observation Template files.")
            return

        self.progress["value"] = 0
        self.root.update()

        try:
            nessus_dfs = []
            total_files = len(self.nessus_files)
            for idx, file_path in enumerate(self.nessus_files):
                df = pd.read_csv(file_path)
                nessus_dfs.append(df)
                self.progress["value"] = 20 * (idx + 1) / total_files
                self.root.update()
            nessus_df = pd.concat(nessus_dfs, ignore_index=True)
        except Exception as e:
            self.status_label.config(text=f"Error: Failed to read Nessus CSV(s): {str(e)}")
            return

        required_columns = ["Risk", "Host", "Protocol", "Port", "Name"]
        available_columns = [col for col in required_columns if col in nessus_df.columns]
        if not all(col in nessus_df.columns for col in ["Host", "Name"]):
            self.status_label.config(text="Error: Required columns (Host, Name) not found in Nessus CSV(s).")
            return

        self.filtered_df = nessus_df[available_columns]
        valid_risks = ["critical", "high", "medium", "low", "info"]
        if "Risk" in self.filtered_df.columns:
            self.filtered_df = self.filtered_df[
                self.filtered_df["Risk"].notna() & 
                self.filtered_df["Risk"].str.lower().isin(valid_risks)
            ]
        if "Protocol" in self.filtered_df.columns:
            self.filtered_df = self.filtered_df[self.filtered_df["Protocol"] != "icmp"]
        self.filtered_df["Name"] = self.filtered_df["Name"].astype(str).apply(self.clean_string)

        for item in self.filtered_tree.get_children():
            self.filtered_tree.delete(item)
        display_columns = ["Risk", "Host", "Protocol", "Port", "Name"]
        for _, row in self.filtered_df[[col for col in display_columns if col in self.filtered_df.columns]].iterrows():
            self.filtered_tree.insert("", "end", values=tuple(row.get(col, "") for col in display_columns))
        rows, cols = self.filtered_df[[col for col in display_columns if col in self.filtered_df.columns]].shape
        total_values = rows * cols
        self.filtered_stats.config(text=f"Total Rows: {rows} | Total Columns: {cols} | Total Values: {total_values}")
        if rows == 0:
            self.status_label.config(text="Warning: No valid rows found in Nessus CSV(s) after filtering.")
            return
        self.progress["value"] = 40
        self.root.update()

        try:
            template_df = pd.read_excel(self.template_file)
            header_row = self.find_header_row(template_df, "S_No")
            if header_row is None:
                self.status_label.config(text="Error: S_No not found in Observation Template.")
                return
            template_df.columns = template_df.iloc[header_row]
            template_df = template_df.iloc[header_row + 1:].reset_index(drop=True)
            template_df["Observation Name"] = template_df["Observation Name"].astype(str).apply(self.clean_string)
            template_df["Severity"] = template_df["Severity"].astype(str).str.strip().str.lower()
        except Exception as e:
            self.status_label.config(text=f"Error: Failed to read Observation Template: {str(e)}")
            return

        self.progress["value"] = 60
        self.root.update()

        output_rows = []
        not_found_rows = []
        matched_names = set()
        total_rows = len(self.filtered_df)
        for idx, nessus_row in enumerate(self.filtered_df.iterrows()):
            match_found = False
            _, nessus_row = nessus_row
            nessus_name = self.clean_string(nessus_row["Name"])
            best_match_row = None

            for _, template_row in template_df.iterrows():
                template_name = self.clean_string(template_row["Observation Name"])
                if nessus_name == template_name:
                    best_match_row = template_row
                    break

            if best_match_row is not None:
                output_row = best_match_row.copy()
                output_row["Affected IP Address"] = nessus_row["Host"]
                output_row["Port Number"] = nessus_row.get("Port", "")
                output_row["Port Type"] = nessus_row.get("Protocol", "")
                output_rows.append(output_row)
                matched_names.add(nessus_name)
                match_found = True

            if not match_found:
                not_found_rows.append(nessus_row)

            self.progress["value"] = 60 + 20 * (idx + 1) / total_rows
            self.root.update()

        self.not_found_df = pd.DataFrame(not_found_rows) if not_found_rows else pd.DataFrame(columns=available_columns)

        if not output_rows:
            self.status_label.config(text="Info: No matches found between Nessus CSV(s) and Observation Template.")
            return

        self.output_df = pd.DataFrame(output_rows)
        self.output_df = self.output_df.sort_values(by="Severity", key=lambda x: x.map({
            "critical": 1, "high": 2, "medium": 3, "low": 4, "info": 5
        }))

        if len(self.output_df) != len(self.filtered_df):
            self.status_label.config(text=f"Warning: Filtered Nessus rows ({len(self.filtered_df)}) do not match output rows ({len(self.output_df)}). Unmatched rows exported to 'Not Found' sheet.")

        self.progress["value"] = 80
        self.root.update()

        for item in self.output_tree.get_children():
            self.output_tree.delete(item)
        if not self.output_df.empty:
            # Exclude the second column (index 1) after S_No for display in the UI
            columns = list(self.output_df.columns)
            if "S_No" in columns:
                columns = ["S_No"] + [col for col in columns if col != "S_No"]
            if len(columns) > 1:
                columns.pop(1)  # Remove the second column (index 1)
            self.output_tree["columns"] = columns
            for col in columns:
                self.output_tree.heading(col, text=col)
                self.output_tree.column(col, width=150, stretch=True)
            for _, row in self.output_df.iterrows():
                self.output_tree.insert("", "end", values=[row.get(col, "") for col in columns])
        
        rows, cols = len(self.output_df), len(columns) if not self.output_df.empty else (0, 0)
        total_values = rows * cols
        self.output_stats.config(text=f"Total Rows: {rows} | Total Columns: {cols} | Total Values: {total_values} | Matches: {rows}")
        self.progress["value"] = 100
        self.root.update()

        self.status_label.config(text="Status: Processing complete.")
        self.export_btn.state(['!disabled'])

    def export_to_excel(self):
        if self.output_df is None:
            self.status_label.config(text="Error: No processed data to export.")
            return

        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_file:
            return

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        colors = {
            "header": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"),
            "critical": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
            "high": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            "low": PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            "info": PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
        }

        # Observation 1 Sheet (Ensure S_No is the first column, remove the second column)
        ws = wb.create_sheet("Observation 1")
        # Reorder columns to ensure S_No is the first column
        if "S_No" in self.output_df.columns:
            columns = ["S_No"] + [col for col in self.output_df.columns if col != "S_No"]
        else:
            columns = ["S_No"] + list(self.output_df.columns)
        # Remove the second column (index 1) which corresponds to column B
        if len(columns) > 1:
            columns.pop(1)  # Remove the second column (index 1)
        headers = columns
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = colors["header"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row in enumerate(self.output_df.itertuples(index=False), 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if header == "S_No":
                    cell.value = row_idx - 1  # Sequential S_No
                else:
                    # Find the value from the row corresponding to the header
                    try:
                        value = self.output_df.iloc[row_idx-2][header]
                        cell.value = value
                    except (KeyError, IndexError):
                        cell.value = ""
                if header == "Observation Name":
                    cell.value = str(cell.value).title()
                if header == "Severity":
                    cell.value = str(cell.value).title()
                    cell.font = Font(bold=True)
                    severity = str(cell.value).lower()
                    if severity in colors:
                        cell.fill = colors[severity]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Filtered Data Sheet
        ws_filtered = wb.create_sheet("Filtered Data")
        filtered_headers = self.filtered_df.columns.tolist()
        for col_idx, header in enumerate(filtered_headers, 1):
            cell = ws_filtered.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = colors["header"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row in enumerate(self.filtered_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_filtered.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        for col in ws_filtered.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_filtered.column_dimensions[column].width = adjusted_width

        # Not Found Sheet
        if not self.not_found_df.empty:
            ws_not_found = wb.create_sheet("Not Found")
            not_found_headers = self.not_found_df.columns.tolist()
            for col_idx, header in enumerate(not_found_headers, 1):
                cell = ws_not_found.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = colors["header"]
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            for row_idx, row in enumerate(self.not_found_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    cell = ws_not_found.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border

            for col in ws_not_found.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws_not_found.column_dimensions[column].width = adjusted_width

        wb.save(output_file)
        self.status_label.config(text=f"Status: Exported to: {os.path.basename(output_file)}")
        self.progress["value"] = 0

if __name__ == "__main__":
    root = tk.Tk()
    app = NessusObservationProcessor(root)
    root.mainloop()