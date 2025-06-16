import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import re
import csv
import pyperclip
from datetime import datetime

class NmapPortViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Open Port Details")
        
        # Create main scrollable frame
        main_canvas = tk.Canvas(root)
        scrollbar = ttk.Scrollbar(root, orient="vertical", command=main_canvas.yview)
        self.scrollable_frame = ttk.Frame(main_canvas)
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        
        main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)
        
        # Grid layout for scrollable main window
        main_canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        root.grid_rowconfigure(0, weight=1)
        root.grid_columnconfigure(0, weight=1)
        
        # Create header with blue theme
        header_frame = tk.Frame(self.scrollable_frame, bg="#0078D7", relief="raised", borderwidth=1)
        header_frame.grid(row=0, column=0, sticky="ew", padx=1, pady=1)
        header_label = tk.Label(header_frame, text="Open Port Details", 
                              bg="#0078D7", fg="white", 
                              font=("Segoe UI", 12, "bold"),
                              pady=5)
        header_label.pack(fill="x")
        
        # Add search frame at the top
        search_frame = ttk.Frame(self.scrollable_frame)
        search_frame.grid(row=1, column=0, sticky="ew", pady=5, padx=5)
        ttk.Label(search_frame, text="Search:").pack(side="left", padx=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.apply_filters())
        ttk.Entry(search_frame, textvariable=self.search_var).pack(side="left", fill="x", expand=True, padx=5)
        
        # Create main frame
        self.main_frame = ttk.Frame(self.scrollable_frame, padding="5")
        self.main_frame.grid(row=2, column=0, sticky="nsew")
        
        # Initialize variables first
        self.sort_column = None
        self.sort_reverse = False
        
        # Initialize filter variables
        self.filter_vars = {}
        for col_name in ["IP Address", "Port Number", "Status", "Service", "Version"]:
            self.filter_vars[col_name] = {
                'var': tk.StringVar(value="All"),
                'values': set(["All"]),
                'combo': None
            }
        
        # Create header with blue theme
        header_frame = tk.Frame(root, bg="#0078D7", relief="raised", borderwidth=1)
        header_frame.grid(row=0, column=0, sticky="ew", padx=1, pady=1)
        header_label = tk.Label(header_frame, text="Open Port Details", 
                              bg="#0078D7", fg="white", 
                              font=("Segoe UI", 12, "bold"),
                              pady=5)
        header_label.pack(fill="x")
        
        # Create main frame
        self.main_frame = ttk.Frame(root, padding="5")
        self.main_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create button frame with status label
        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.grid(row=0, column=0, columnspan=2, pady=5)
        
        # Create status label first
        self.status_label = ttk.Label(self.button_frame, text="Ready")
        self.status_label.pack(side="right", padx=10)
        
        # Add search bar in the top corner
        search_frame = ttk.Frame(self.button_frame)
        search_frame.pack(side="right", padx=10, before=self.status_label)
        ttk.Label(search_frame, text="Search:").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.apply_filters())
        ttk.Entry(search_frame, textvariable=self.search_var, width=30).pack(side="left", padx=5)
        
        # Create buttons with icons
        ttk.Button(self.button_frame, text="Upload TCP", 
                  command=self.tcp_upload).pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Upload UDP", 
                  command=self.udp_upload).pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Copy to Clipboard", 
                  command=self.copy_to_clipboard).pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Export to Excel", 
                  command=self.export_to_excel).pack(side="left", padx=5)
        ttk.Button(self.button_frame, text="Clear All", 
                  command=self.clear_data).pack(side="left", padx=5)
        
        # Create filter frame with checkboxes
        filter_frame = ttk.LabelFrame(self.main_frame, text="Filters", padding="5")
        filter_frame.grid(row=1, column=0, sticky="ew", pady=5)
        
        # Create filter sections with checkboxes
        for i, col_name in enumerate(["IP Address", "Port Number", "Status", "Service", "Version"]):
            frame = ttk.LabelFrame(filter_frame, text=col_name)
            frame.grid(row=0, column=i, padx=5, pady=5, sticky="nsew")
            
            # Create canvas and scrollbar for checkboxes
            canvas = tk.Canvas(frame, height=150)
            scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
            checkbox_frame = ttk.Frame(canvas)
            
            canvas.configure(yscrollcommand=scrollbar.set)
            
            # Pack widgets
            scrollbar.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)
            
            # Create window for checkboxes
            canvas_frame = canvas.create_window((0, 0), window=checkbox_frame, anchor="nw")
            
            # Store checkbox variables
            self.filter_vars[col_name] = {
                'vars': {},
                'frame': checkbox_frame,
                'canvas': canvas,
                'values': set(["All"])
            }
            
            # Add select all checkbox
            all_var = tk.BooleanVar(value=True)
            all_cb = ttk.Checkbutton(checkbox_frame, text="Select All", 
                                   variable=all_var, 
                                   command=lambda cn=col_name: self.toggle_all(cn))
            all_cb.pack(anchor="w")
            self.filter_vars[col_name]['all_var'] = all_var
            
            # Configure canvas scrolling
            checkbox_frame.bind("<Configure>", lambda e, c=canvas: self.configure_scroll_region(c))
            canvas.bind("<Configure>", lambda e, c=canvas, f=canvas_frame: self.configure_canvas_window(e, c, f))
            
            # Enable mousewheel scrolling
            canvas.bind_all("<MouseWheel>", lambda e, c=canvas: self.on_mousewheel(e, c))
        
        # Add Select All/None buttons
        button_frame = ttk.Frame(filter_frame)
        button_frame.grid(row=1, column=0, columnspan=5, pady=5)
        ttk.Button(button_frame, text="Select All", command=self.select_all_filters).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Select None", command=self.unselect_all_filters).pack(side="left", padx=5)
        
        # Create TCP and UDP frames
        self.tcp_frame = ttk.LabelFrame(self.main_frame, text="TCP Open Port Details", padding="5")
        self.tcp_frame.grid(row=2, column=0, sticky="nsew", pady=5)
        
        self.udp_frame = ttk.LabelFrame(self.main_frame, text="UDP Open Port Details", padding="5")
        self.udp_frame.grid(row=3, column=0, sticky="nsew", pady=5)
        
        # Create TCP and UDP trees
        self.tcp_tree = self.create_treeview(self.tcp_frame)
        self.udp_tree = self.create_treeview(self.udp_frame)
        
        # Status label
        self.status_label = ttk.Label(self.button_frame, text="Ready")
        self.status_label.pack(side="right", padx=10)
        
        # Create and style Treeview
        style = ttk.Style()
        style.configure("Treeview.Heading", font=('Segoe UI', 9, 'bold'))
        style.configure("Treeview", font=('Segoe UI', 9), rowheight=25)

    def create_treeview(self, parent):
        tree = ttk.Treeview(parent, 
                       columns=("IP", "Port", "Status", "Service", "Version"),
                       show="headings")
        
        # Configure headings and columns with text wrapping
        columns = {
            "IP": ("IP Address", 120),
            "Port": ("Port Number", 100),
            "Status": ("Status", 80),
            "Service": ("Service", 150),
            "Version": ("Version", 400)  # Increased width for version
        }
        
        for col, (heading, width) in columns.items():
            tree.heading(col, text=heading)
            tree.column(col, width=width, anchor="w", stretch=True)  # Added stretch and anchor
        
        # Add scrollbars
        y_scrollbar = ttk.Scrollbar(parent, orient="vertical", command=tree.yview)
        x_scrollbar = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # Grid layout
        tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # Configure grid weights
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        
        return tree

    def copy_to_clipboard(self):
        data = []
        for tree in [self.tcp_tree, self.udp_tree]:
            for item in tree.get_children():
                values = tree.item(item)['values']
                data.append('\t'.join(str(v) for v in values))
        pyperclip.copy('\n'.join(data))

    def write_excel_section(self, ws, title, tree, start_row):
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Define styles
        blue_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
        row_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Title row with updated text
        ws.merge_cells(f'A{start_row}:E{start_row}')
        title_cell = ws[f'A{start_row}']
        title_cell.value = title + " Open Port Details"  # Updated title text
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = blue_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = ["IP Address", "Port Number", "Status", "Service", "Version"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row + 1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = thin_border
        
        # Data rows - modified to respect filters
        current_row = start_row + 2
        for item in tree.get_children():
            # Skip hidden items (filtered out)
            if 'hidden' in tree.item(item, 'tags'):
                continue
                
            values = tree.item(item)['values']
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                cell.fill = row_fill
            current_row += 1
        
        # Set column widths with better spacing
        ws.column_dimensions['A'].width = 20  # IP
        ws.column_dimensions['B'].width = 15  # Port
        ws.column_dimensions['C'].width = 12  # Status
        ws.column_dimensions['D'].width = 25  # Service
        ws.column_dimensions['E'].width = 60  # Version
        
        return current_row

    def configure_scroll_region(self, canvas):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def configure_canvas_window(self, event, canvas, frame):
        canvas.itemconfig(frame, width=event.width)

    def on_mousewheel(self, event, canvas):
        if canvas.winfo_height() < canvas.bbox("all")[3]:
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    def toggle_all(self, col_name):
        all_selected = self.filter_vars[col_name]['all_var'].get()
        for var in self.filter_vars[col_name]['vars'].values():
            var.set(all_selected)
        self.apply_filters()

    def update_filter_values(self):
        # Clear existing checkboxes
        for col_name in self.filter_vars:
            for widget in self.filter_vars[col_name]['frame'].winfo_children():
                if widget.winfo_class() == 'TCheckbutton' and widget.cget('text') != "Select All":
                    widget.destroy()
            self.filter_vars[col_name]['vars'].clear()
            self.filter_vars[col_name]['values'] = set()

        # Collect unique values
        for tree in [self.tcp_tree, self.udp_tree]:
            for item in tree.get_children():
                values = tree.item(item)['values']
                for i, col_name in enumerate(["IP Address", "Port Number", "Status", "Service", "Version"]):
                    if values[i]:
                        self.filter_vars[col_name]['values'].add(str(values[i]))

        # Create checkboxes for unique values
        for col_name in self.filter_vars:
            sorted_values = sorted(list(self.filter_vars[col_name]['values']))
            for value in sorted_values:
                var = tk.BooleanVar(value=True)
                cb = ttk.Checkbutton(self.filter_vars[col_name]['frame'], 
                                   text=value, 
                                   variable=var,
                                   command=self.apply_filters)
                cb.pack(anchor="w")
                self.filter_vars[col_name]['vars'][value] = var

    def apply_filters(self, event=None):
        search_term = self.search_var.get().lower()
        
        for tree in [self.tcp_tree, self.udp_tree]:
            for item in tree.get_children():
                values = tree.item(item)['values']
                show_item = True
                
                # Check checkbox filters
                for i, col_name in enumerate(["IP Address", "Port Number", "Status", "Service", "Version"]):
                    value = str(values[i])
                    if value in self.filter_vars[col_name]['vars']:
                        if not self.filter_vars[col_name]['vars'][value].get():
                            show_item = False
                            break
                
                # Apply search filter if item passed checkbox filters
                if show_item and search_term:
                    show_item = any(search_term in str(v).lower() for v in values)
                
                # Apply visibility
                if show_item:
                    tree.item(item, tags=())
                else:
                    tree.item(item, tags=('hidden',))
            
            tree.tag_configure('hidden', background='gray90')

    def select_all_filters(self):
        for col_name in self.filter_vars:
            self.filter_vars[col_name]['var'].set("All")
        self.apply_filters()

    def unselect_all_filters(self):
        # Set first value after "All" for each filter
        for col_name in self.filter_vars:
            values = self.filter_vars[col_name]['combo']['values']
            if len(values) > 1:
                self.filter_vars[col_name]['var'].set(values[1])
        self.apply_filters()

    def search_treeview(self, *args):
        search_term = self.search_var.get().lower()
        for tree in [self.tcp_tree, self.udp_tree]:
            for item in tree.get_children():
                if search_term:
                    values = [str(v).lower() for v in tree.item(item)['values']]
                    if any(search_term in v for v in values):
                        tree.item(item, tags=('match',))
                    else:
                        tree.item(item, tags=('nomatch',))
                else:
                    tree.item(item, tags=())
            
            tree.tag_configure('match', background='lightblue')
            tree.tag_configure('nomatch', background='white')

    def sort_treeview(self, tree, col):
        items = [(tree.set(item, col), item) for item in tree.get_children('')]
        
        # Change sort order if clicking the same column
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_reverse = False
            self.sort_column = col
        
        items.sort(reverse=self.sort_reverse)
        for index, (val, item) in enumerate(items):
            tree.move(item, '', index)

    def clear_data(self):
        if messagebox.askyesno("Clear Data", "Are you sure you want to clear all data?"):
            self.tcp_tree.delete(*self.tcp_tree.get_children())
            self.udp_tree.delete(*self.udp_tree.get_children())
            self.status_label.config(text="Data cleared")

    def update_status(self, message):
        self.status_label.config(text=message)
        self.root.after(3000, lambda: self.status_label.config(text="Ready"))

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select Nmap Output Files",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Process each selected file
        for file_path in files:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                results = self.parse_nmap_output(content)
                for result in results:
                    self.tree.insert("", tk.END, values=result)

    def parse_nmap_output(self, file_content):
        results = []
        current_ip = ""
        
        for line in file_content.split('\n'):
            ip_match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
            if ip_match:
                current_ip = ip_match.group(1)
            
            port_match = re.match(r'(\d+)/(tcp|udp)\s+(\w+)\s+(.+?)\s+(.*)', line)
            if port_match and 'open' in line:
                port, proto, status, service, version = port_match.groups()
                # Clean up version information
                version = re.sub(r'syn-ack ttl \d+', '', version).strip()
                version = re.sub(r'ttl \d+', '', version).strip()
                # Clean up service name
                service = service.replace('?', '').strip()
                if not service:
                    service = "unknown"
                if status == 'open':
                    results.append((current_ip, f"{port}/{proto}", status, service, version))
        
        return results

    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if file_path:
            try:
                import openpyxl
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Port Details"
                
                # Write TCP section
                current_row = 1
                self.write_excel_section(ws, "TCP Ports", self.tcp_tree, current_row)
                
                # Add spacing between sections
                current_row = ws.max_row + 2
                
                # Write UDP section
                self.write_excel_section(ws, "UDP Ports", self.udp_tree, current_row)
                
                wb.save(file_path)
                self.update_status("Report exported successfully")
                
            except ImportError:
                messagebox.showerror("Error", "Please install openpyxl: pip install openpyxl")

    def tcp_upload(self):
        files = filedialog.askopenfilenames(
            title="Select TCP Nmap Files",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        self.process_files(files, 'tcp')
    
    def udp_upload(self):
        files = filedialog.askopenfilenames(
            title="Select UDP Nmap Files",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        self.process_files(files, 'udp')
    
    def process_files(self, files, protocol):
        if not files:
            return
        
        target_tree = self.tcp_tree if protocol == 'tcp' else self.udp_tree
        
        for file_path in files:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
                results = self.parse_nmap_output(content)
                for result in results:
                    if protocol in result[1].lower():
                        target_tree.insert("", tk.END, values=result)
        
        self.update_status(f"{protocol.upper()} data loaded")
        self.update_filter_values()  # Call update_filter_values after loading data

if __name__ == "__main__":
    root = tk.Tk()
    app = NmapPortViewer(root)
    root.mainloop()